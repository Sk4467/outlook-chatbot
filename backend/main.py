import os
import io
import hashlib
from datetime import datetime
from typing import List, Optional
import fitz  # PyMuPDF
from fastapi import FastAPI, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook


import chromadb
from chromadb.config import Settings

import google.generativeai as genai

# =========================

# Setup (runtime only)

# =========================

GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")
if not GOOGLE_API_KEY:
    raise RuntimeError("Set GOOGLE_API_KEY environment variable.")
genai.configure(api_key=GOOGLE_API_KEY)

# Chroma in-memory client (no persistence)

client = chromadb.Client(Settings(anonymized_telemetry=False))
collection = client.create_collection(name="emails", metadata={"hnsw:space": "cosine"})

# In-session content hash set for simple idempotency

seen_hashes = set()

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # dev only
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Gmail IMAP support
from .gmail_imap import fetch_emails as gmail_fetch_emails

# =========================

# Utilities

# =========================

def stable_email_id(email_subject: str, email_from: str, received_at: str) -> str:
    h = hashlib.md5(f"{email_subject}|{email_from}|{received_at}".encode("utf-8")).hexdigest()
    return h[:12]

def sha1_bytes(b: bytes) -> str:
    return hashlib.sha1(b).hexdigest()

def chunk_text(text: str, max_words: int = 900, overlap: int = 80) -> List[str]:
    words = text.split()
    if not words:
        return []
    chunks = []
    start = 0
    while start < len(words):
        end = min(len(words), start + max_words)
        chunk_words = words[start:end]
        chunks.append(" ".join(chunk_words))
        if end == len(words):
            break
        start = max(0, end - overlap)
    return chunks

def embed_text(text: str) -> List[float]:
    resp = genai.embed_content(model="models/text-embedding-004", content=text)
    return resp["embedding"]

def embed_texts(texts: List[str]) -> List[List[float]]:
    embeddings = []
    for t in texts:
        embeddings.append(embed_text(t))
    return embeddings

def parse_pdf_bytes(pdf_bytes: bytes) -> List[dict]:
    
    docs = []
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text("text")
        docs.append({
            "text": text or "",
            "page": page_num + 1,
        })
    doc.close()
    return docs

def parse_xlsx_bytes(xlsx_bytes: bytes) -> List[dict]:
    
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    docs = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        for r_idx, row in enumerate(rows[1:], start=2):
            cells = []
            for c_idx, val in enumerate(row):
                header = headers[c_idx] if c_idx < len(headers) else f"Col{c_idx+1}"
                val_str = "" if val is None else str(val)
                cells.append(f"{header}: {val_str}")
            row_text = f"Sheet: {sheet.title}, Row {r_idx} | " + " | ".join(cells)
            docs.append({
                "text": row_text,
                "sheet": sheet.title,
                "row": r_idx,
            })
    return docs

def build_prompt(question: str, contexts: List[dict]) -> str:
    header = (
        "You are a helpful assistant answering questions using ONLY the provided context. "
        "If the answer isn't in the context, say you don't know. Cite sources with page/sheet info.\n\n"
    )
    ctx_blocks = []
    for i, c in enumerate(contexts, start=1):
        src = c.get("source", "unknown")
        loc = c.get("loc", "")
        text = c.get("text", "")
        ctx_blocks.append(f"[{i}] Source: {src} {loc}\n{text}\n")
    ctx_str = "\n".join(ctx_blocks)
    return f"{header}Context:\n{ctx_str}\nQuestion: {question}\nAnswer:"

def gemini_answer(prompt: str) -> str:
    model = genai.GenerativeModel("gemini-1.5-flash")
    resp = model.generate_content(prompt)
    return (resp.text or "").strip()

# =========================

# API Models

# =========================

class ChatRequest(BaseModel):
    query: str
    k: Optional[int] = 6

# =========================

# Endpoints

# =========================

@app.post("/ingest")
async def ingest(
    emailSubject: str = Form(...),
    emailFrom: str = Form(...),
    receivedAt: str = Form(...),
    bodyText: str = Form(""),
    attachments: List[UploadFile] = File(None),
):
    email_id = stable_email_id(emailSubject, emailFrom, receivedAt)

    docs_texts = []
    docs_metas = []
    ids = []

    # Ingest email body
    if bodyText and bodyText.strip():
        for idx, chunk in enumerate(chunk_text(bodyText)):
            b = chunk.encode("utf-8")
            chash = sha1_bytes(b)
            if chash in seen_hashes:
                continue
            seen_hashes.add(chash)
            docs_texts.append(chunk)
            docs_metas.append({
                "type": "email-body",
                "emailId": email_id,
                "subject": emailSubject,
                "from": emailFrom,
                "receivedAt": receivedAt,
                "provenance": "email body",
                "source": f"Email: {emailSubject}",
                "loc": "",
            })
            ids.append(f"{email_id}-body-{idx}")

    # Ingest attachments
    count_attachments = 0
    if attachments:
        for f in attachments:
            content = await f.read()
            ct = (f.content_type or "").lower()
            base_meta = {
                "type": "attachment",
                "emailId": email_id,
                "subject": emailSubject,
                "from": emailFrom,
                "receivedAt": receivedAt,
                "filename": f.filename,
            }
            parsed_chunks = []
            if "pdf" in ct or f.filename.lower().endswith(".pdf"):
                pages = parse_pdf_bytes(content)
                for p in pages:
                    text = p["text"]
                    for idx, chunk in enumerate(chunk_text(text)):
                        parsed_chunks.append({
                            "text": chunk,
                            "source": f.filename,
                            "loc": f"(page {p['page']})",
                        })
            elif "spreadsheetml.sheet" in ct or f.filename.lower().endswith(".xlsx"):
                rows = parse_xlsx_bytes(content)
                for r in rows:
                    parsed_chunks.append({
                        "text": r["text"],
                        "source": f.filename,
                        "loc": f"(sheet {r['sheet']}, row {r['row']})",
                    })
            else:
                # skip unsupported types in prototype
                continue

            if parsed_chunks:
                count_attachments += 1

            for idx, pc in enumerate(parsed_chunks):
                b = pc["text"].encode("utf-8")
                chash = sha1_bytes(b)
                if chash in seen_hashes:
                    continue
                seen_hashes.add(chash)

                docs_texts.append(pc["text"])
                meta = {
                    **base_meta,
                    "provenance": f"attachment {f.filename}",
                    "source": pc["source"],
                    "loc": pc["loc"],
                }
                docs_metas.append(meta)
                ids.append(f"{email_id}-{f.filename}-{idx}")

    if not docs_texts:
        return {"chunks": 0, "attachments": count_attachments}

    embeddings = embed_texts(docs_texts)
    collection.add(documents=docs_texts, embeddings=embeddings, metadatas=docs_metas, ids=ids)
    return {"chunks": len(docs_texts), "attachments": count_attachments}

@app.post("/chat")
async def chat(req: ChatRequest):
    if req.k is None or req.k <= 0:
        req.k = 6
    q_emb = embed_text(req.query)
    res = collection.query(query_embeddings=[q_emb], n_results=req.k)
    docs = res.get("documents", [[]])[0]
    metas = res.get("metadatas", [[]])[0]

    contexts = []
    sources = []
    for text, m in zip(docs, metas):
        src = m.get("source") or m.get("filename") or "unknown"
        loc = m.get("loc") or ""
        contexts.append({"text": text, "source": src, "loc": loc})
        sources.append({
            "subject": m.get("subject"),
            "filename": m.get("filename"),
        })

    prompt = build_prompt(req.query, contexts)
    answer = gemini_answer(prompt)
    return {"answer": answer, "sources": sources}


# =========================
# Gmail IMAP Endpoints (runtime only)
# =========================

@app.get("/gmail/messages")
def gmail_messages(
    user: int = Query(1, description="User slot: 1 or 2"),
    n: int = Query(10, description="Last N emails"),
):
    """
    Reads emails for one of two configured users.
    Configure environment variables:
      - GMAIL_USER1, GMAIL_PASS1
      - GMAIL_USER2, GMAIL_PASS2
    """
    if user not in (1, 2):
        return {"value": []}
    email_account = os.getenv(f"GMAIL_USER{user}", "")
    email_password = os.getenv(f"GMAIL_PASS{user}", "")
    if not email_account or not email_password:
        return {"error": f"Missing GMAIL_USER{user}/GMAIL_PASS{user} env vars"}
    msgs = gmail_fetch_emails(email_account, email_password, n=n)
    return {"value": msgs}


@app.get("/gmail/users")
def gmail_users():
    users = []
    for slot in (1, 2):
        email_user = os.getenv(f"GMAIL_USER{slot}")
        users.append({
            "slot": slot,
            "email": email_user or None,
            "configured": bool(email_user),
        })
    return {"users": users}


@app.get("/gmail/messages/all")
def gmail_messages_all(n: int = Query(10, description="Last N emails")):
    out = {}
    for slot in (1, 2):
        email_account = os.getenv(f"GMAIL_USER{slot}", "")
        email_password = os.getenv(f"GMAIL_PASS{slot}", "")
        if email_account and email_password:
            out[str(slot)] = gmail_fetch_emails(email_account, email_password, n=n)
        else:
            out[str(slot)] = {"error": f"Missing GMAIL_USER{slot}/GMAIL_PASS{slot}"}
    return out
