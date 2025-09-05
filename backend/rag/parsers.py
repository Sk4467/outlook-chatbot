import io
from typing import List, Dict, Any, Tuple
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import fitz  # PyMuPDF

def html_to_text(html: str) -> str:
    return BeautifulSoup(html or "", "html.parser").get_text(separator=" ", strip=True)

def chunk_text(text: str, max_words: int = 900, overlap: int = 80) -> List[str]:
    words = (text or "").split()
    if not words:
        return []
    chunks, start = [], 0
    while start < len(words):
        end = min(len(words), start + max_words)
        chunks.append(" ".join(words[start:end]))
        if end == len(words): break
        start = max(0, end - overlap)
    return chunks

def parse_pdf_bytes(pdf_bytes: bytes) -> List[Dict[str, Any]]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    out: List[Dict[str, Any]] = []
    for i in range(len(doc)):
        page = doc[i]
        text = page.get_text("text") or ""
        
        # Enhanced text extraction with structure preservation
        # Try to get more structured text if available
        try:
            # Get text with layout information
            blocks = page.get_text("dict")
            structured_text = _extract_structured_text(blocks)
            if structured_text and len(structured_text) > len(text):
                text = structured_text
        except:
            # Fallback to simple text extraction
            pass
        
        # Clean and normalize text
        text = _clean_pdf_text(text)
        
        out.append({
            "text": text, 
            "page": i + 1,
            "char_count": len(text),
            "word_count": len(text.split())
        })
    doc.close()
    return out

def _extract_structured_text(blocks_dict: Dict[str, Any]) -> str:
    """Extract text while preserving some structure from PDF blocks"""
    text_parts = []
    
    if "blocks" in blocks_dict:
        for block in blocks_dict["blocks"]:
            if "lines" in block:
                block_text = []
                for line in block["lines"]:
                    if "spans" in line:
                        line_text = ""
                        for span in line["spans"]:
                            if "text" in span:
                                line_text += span["text"]
                        if line_text.strip():
                            block_text.append(line_text.strip())
                
                if block_text:
                    # Join lines within a block
                    text_parts.append(" ".join(block_text))
    
    # Join blocks with double newlines to preserve paragraph structure
    return "\n\n".join(text_parts)

def _clean_pdf_text(text: str) -> str:
    """Clean and normalize PDF text"""
    import re
    
    # Remove excessive whitespace
    text = re.sub(r'\s+', ' ', text)
    
    # Remove page headers/footers patterns (common patterns)
    text = re.sub(r'Page \d+ of \d+', '', text)
    text = re.sub(r'^\d+\s*$', '', text, flags=re.MULTILINE)
    
    # Fix common PDF extraction issues
    text = text.replace('\uf0b7', '•')  # Replace bullet point character
    text = text.replace('\u2022', '•')  # Another bullet point
    text = text.replace('\u2013', '-')  # En dash
    text = text.replace('\u2014', '--')  # Em dash
    text = text.replace('\u201c', '"')  # Left double quote
    text = text.replace('\u201d', '"')  # Right double quote
    text = text.replace('\u2018', "'")  # Left single quote
    text = text.replace('\u2019', "'")  # Right single quote
    
    # Remove extra spaces and normalize
    text = re.sub(r' +', ' ', text)
    text = text.strip()
    
    return text

def xlsx_summary_from_bytes(xlsx_bytes: bytes, sample_rows: int = 50) -> List[Dict[str, Any]]:
    """
    Return summary entries suitable for an index (NOT full-row embeddings).
    Each entry gets sheet name, columns, row_count, and a short sample preview.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    summaries: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [(str(h).strip() if h is not None else "") for h in (rows[0] or [])]
        data_rows = rows[1:]
        row_count = len(data_rows)
        sample = data_rows[:sample_rows]
        # Render a compact preview (first few rows)
        preview_lines = []
        for r in sample:
            pairs = []
            for i, val in enumerate(r):
                col = headers[i] if i < len(headers) else f"Col{i+1}"
                pairs.append(f"{col}: {'' if val is None else str(val)}")
            preview_lines.append(" | ".join(pairs))
        text = f"Sheet: {ws.title}\nColumns: {', '.join(h for h in headers if h)}\nRows: {row_count}\nPreview:\n" + "\n".join(preview_lines[:5])
        summaries.append({
            "sheet": ws.title,
            "columns": headers,
            "row_count": row_count,
            "text": text,
        })
    return summaries